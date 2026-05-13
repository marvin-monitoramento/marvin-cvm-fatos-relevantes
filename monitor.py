"""
Marvin - Monitor de Fatos Relevantes CVM (Cloud Edition para GitHub Actions)
=============================================================================

Versao cloud do monitor de Fatos Relevantes. Roda em GitHub Actions diariamente.
Diferencas vs versao local:
  - Sem pywin32/Outlook (nao funciona em servidor Linux)
  - Notificacao via ClickUp (cria task na lista de alertas, ClickUp envia o e-mail)
  - Estado (historico/state) commitado no proprio repo apos cada run
  - Variaveis de ambiente para secrets

Variaveis de ambiente esperadas:
  CLICKUP_API_TOKEN          - token pessoal do ClickUp (pk_xxx)
  CLICKUP_LIST_ID_ALERTAS    - ID da lista "Alertas CVM Fatos Relevantes"
  CLICKUP_ASSIGNEE_IDS       - ID(s) do(s) usuario(s) responsavel(eis), separados por ","

Uso:
  python monitor.py                # config padrao (lookback 10 dias)
  python monitor.py --lookback 14  # janela maior
  python monitor.py --dry-run      # nao cria task no ClickUp

Saidas:
  historico/historico.csv          - fonte-verdade auditavel (append-only)
  historico/historico_eventos.xlsx - planilha formatada
  data/seen_events.json            - controle de deduplicacao
  data/last_run.json               - metadata do ultimo run

Requer: Python 3.10+, pandas, requests, openpyxl
"""

from __future__ import annotations
import argparse, hashlib, json, logging, os, sys
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import requests

# ---------------------------------------------------------------------------
# Configuracao
# ---------------------------------------------------------------------------

CATEGORIAS_RELEVANTES = {
    "Fato Relevante", "Comunicado ao Mercado", "Aviso aos Acionistas",
    "Calendário de Eventos Corporativos", "Assembleia",
}

TERMOS_CRITICOS = [
    "recuperação judicial", "recuperacao judicial", "RJ",
    "falência", "falencia", "covenant", "inadimplemento", "default",
    "auditor", "auditoria independente", "rating", "agência de classificação",
    "penhora", "arresto", "medida cautelar", "controle acionário",
    "cisão", "fusão", "demissão", "renúncia", "afastamento",
    "ação judicial", "processo", "CARF", "AGU", "debêntures", "dívida",
]

DEFAULT_LOOKBACK_DAYS = 10
HTTP_TIMEOUT_SECONDS = 60
USER_AGENT = "Marvin-CVM-Monitor-Cloud/0.3 (risco@marvin.com.br)"

# Secrets vindos do ambiente
CLICKUP_API_TOKEN = os.environ.get("CLICKUP_API_TOKEN", "")
CLICKUP_LIST_ID_ALERTAS = os.environ.get("CLICKUP_LIST_ID_ALERTAS", "")
CLICKUP_ASSIGNEE_IDS = os.environ.get("CLICKUP_ASSIGNEE_IDS", "")

BASE_DIR = Path(__file__).resolve().parent
WATCHLIST_PATH = BASE_DIR / "watchlist.csv"
STATE_PATH = BASE_DIR / "data" / "seen_events.json"
LAST_RUN_PATH = BASE_DIR / "data" / "last_run.json"
HISTORICO_DIR = BASE_DIR / "historico"
HISTORICO_CSV = HISTORICO_DIR / "historico.csv"
HISTORICO_XLSX = HISTORICO_DIR / "historico_eventos.xlsx"

COLUNAS_HISTORICO = [
    "data_coleta", "data_entrega", "empresa", "cnpj",
    "categoria", "tipo", "assunto", "materialidade",
    "termos_encontrados", "link", "fingerprint",
]

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s")
log = logging.getLogger("marvin-cvm-cloud")


@dataclass
class Evento:
    cnpj: str
    empresa: str
    data_entrega: str
    categoria: str
    tipo: str
    assunto: str
    link: str
    materialidade: str
    termos_encontrados: list

    @property
    def fingerprint(self) -> str:
        raw = f"{self.cnpj}|{self.data_entrega}|{self.categoria}|{self.assunto}"
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def carregar_watchlist():
    if not WATCHLIST_PATH.exists():
        raise FileNotFoundError(f"Watchlist nao encontrada em {WATCHLIST_PATH}")
    df = pd.read_csv(WATCHLIST_PATH, dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]
    if "cnpj" not in df.columns:
        raise ValueError("Watchlist precisa da coluna 'cnpj'")
    df["cnpj_clean"] = df["cnpj"].str.replace(r"[^\d]", "", regex=True)
    df = df[df["cnpj_clean"].str.len() == 14].copy()
    df["cnpj_raiz"] = df["cnpj_clean"].str[:8]
    log.info("Watchlist carregada: %d CNPJs validos (%d raizes unicas)",
             len(df), df["cnpj_raiz"].nunique())
    return df


def baixar_ipe(year=None):
    """Tenta varias combinacoes de URL (csv/zip, ano atual/anterior)."""
    from io import BytesIO
    import zipfile

    anos = [year] if year else [datetime.now().year, datetime.now().year - 1]
    sufixos = [".csv", ".zip"]
    ultima_exc = None

    for ano in anos:
        for suf in sufixos:
            url = f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/IPE/DADOS/ipe_cia_aberta_{ano}{suf}"
            log.info("Tentando: %s", url)
            try:
                resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=HTTP_TIMEOUT_SECONDS)
                resp.raise_for_status()
                conteudo = resp.content
                if suf == ".zip":
                    with zipfile.ZipFile(BytesIO(conteudo)) as zf:
                        nome_csv = [n for n in zf.namelist() if n.lower().endswith(".csv")][0]
                        conteudo = zf.read(nome_csv)
                df = pd.read_csv(BytesIO(conteudo), sep=";", encoding="latin-1",
                                 dtype=str, on_bad_lines="warn")
                log.info("IPE %d carregado: %d linhas", ano, len(df))
                df = df.rename(columns={c: c.strip() for c in df.columns})
                if "Data_Entrega" not in df.columns:
                    for alt in ("Data_Recebimento", "Data_Publicacao", "Data"):
                        if alt in df.columns:
                            df = df.rename(columns={alt: "Data_Entrega"})
                            break
                df["cnpj_clean"] = df["CNPJ_Companhia"].fillna("").str.replace(r"[^\d]", "", regex=True)
                df["cnpj_raiz"] = df["cnpj_clean"].str[:8]
                df["Data_Entrega"] = pd.to_datetime(df["Data_Entrega"], errors="coerce")
                return df
            except requests.HTTPError as exc:
                if exc.response.status_code == 404:
                    log.info("  -> 404, tentando proxima")
                    ultima_exc = exc
                    continue
                raise
            except Exception as exc:
                ultima_exc = exc
                log.warning("  -> falhou: %s", exc)
                continue

    raise RuntimeError(f"Nenhuma URL da CVM funcionou. Ultima: {ultima_exc}")


def classificar(row):
    assunto = str(row.get("Assunto", "")).lower()
    categoria = str(row.get("Categoria", ""))
    encontrados = [t for t in TERMOS_CRITICOS if t.lower() in assunto]
    if categoria == "Fato Relevante":
        base = "alta"
    elif categoria in {"Comunicado ao Mercado", "Assembleia"}:
        base = "média"
    else:
        base = "baixa"
    if encontrados and base != "alta":
        base = "alta"
    return base, encontrados


def filtrar_eventos(ipe, watchlist, lookback_days):
    cutoff = datetime.now() - timedelta(days=lookback_days)
    mascara = (
        ipe["cnpj_raiz"].isin(watchlist["cnpj_raiz"])
        & ipe["Categoria"].isin(CATEGORIAS_RELEVANTES)
        & (ipe["Data_Entrega"] >= cutoff)
    )
    filtrado = ipe.loc[mascara].sort_values("Data_Entrega", ascending=False)
    eventos = []
    for _, row in filtrado.iterrows():
        materialidade, termos = classificar(row)
        eventos.append(Evento(
            cnpj=row["cnpj_clean"],
            empresa=str(row.get("Nome_Companhia", "")),
            data_entrega=row["Data_Entrega"].isoformat() if pd.notna(row["Data_Entrega"]) else "",
            categoria=str(row.get("Categoria", "")),
            tipo=str(row.get("Tipo", "")),
            assunto=str(row.get("Assunto", "")),
            link=str(row.get("Link_Download", "")),
            materialidade=materialidade,
            termos_encontrados=termos,
        ))
    return eventos


def carregar_state():
    if not STATE_PATH.exists():
        return set()
    try:
        return set(json.loads(STATE_PATH.read_text(encoding="utf-8")))
    except Exception:
        return set()


def salvar_state(fps):
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(sorted(fps), ensure_ascii=False, indent=2), encoding="utf-8")


def persistir_historico(eventos):
    HISTORICO_DIR.mkdir(parents=True, exist_ok=True)
    data_coleta = datetime.now().isoformat(timespec="seconds")
    if HISTORICO_CSV.exists():
        df_hist = pd.read_csv(HISTORICO_CSV, dtype=str, keep_default_na=False)
    else:
        df_hist = pd.DataFrame(columns=COLUNAS_HISTORICO)
    fps_exist = set(df_hist["fingerprint"]) if "fingerprint" in df_hist.columns else set()
    novos = []
    for e in eventos:
        if e.fingerprint in fps_exist:
            continue
        novos.append({
            "data_coleta": data_coleta, "data_entrega": e.data_entrega,
            "empresa": e.empresa, "cnpj": e.cnpj, "categoria": e.categoria,
            "tipo": e.tipo, "assunto": e.assunto, "materialidade": e.materialidade,
            "termos_encontrados": "|".join(e.termos_encontrados),
            "link": e.link, "fingerprint": e.fingerprint,
        })
    df_novos = pd.DataFrame(novos, columns=COLUNAS_HISTORICO)
    df_final = pd.concat([df_hist, df_novos], ignore_index=True)
    df_final["_sort"] = pd.to_datetime(df_final["data_entrega"], errors="coerce")
    df_final = df_final.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    df_final.to_csv(HISTORICO_CSV, index=False, encoding="utf-8-sig")
    log.info("Historico: %d total | %d novos neste run", len(df_final), len(df_novos))
    try:
        gerar_xlsx(df_final, HISTORICO_XLSX)
    except Exception as exc:
        log.error("Falha xlsx: %s", exc)
    return df_final


def gerar_xlsx(df, path):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Eventos", index=False)
        if not df.empty:
            resumo = (df.pivot_table(index="empresa", columns="materialidade",
                values="fingerprint", aggfunc="count", fill_value=0)
                .reindex(columns=["alta", "média", "baixa"], fill_value=0))
            resumo["total"] = resumo.sum(axis=1)
            resumo = resumo.sort_values("alta", ascending=False)
            resumo.to_excel(writer, sheet_name="Resumo")
        wb = writer.book
        ws = wb["Eventos"]
        hfill = PatternFill("solid", fgColor="1F3864")
        hfont = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        widths = {"data_coleta": 18, "data_entrega": 20, "empresa": 32, "cnpj": 16,
            "categoria": 22, "tipo": 22, "assunto": 60, "materialidade": 14,
            "termos_encontrados": 30, "link": 45, "fingerprint": 18}
        for idx, col in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)
        fa = PatternFill("solid", fgColor="FDE2E2")
        fm = PatternFill("solid", fgColor="FFF4CE")
        col_mat = df.columns.get_loc("materialidade") + 1 if "materialidade" in df.columns else None
        if col_mat:
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_mat).value
                f = fa if v == "alta" else fm if v == "média" else None
                if f:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = f
        ws.auto_filter.ref = ws.dimensions
        if "Resumo" in wb.sheetnames:
            wr = wb["Resumo"]
            for cell in wr[1]:
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal="center")
            wr.column_dimensions["A"].width = 35
            for ci in range(2, wr.max_column + 1):
                wr.column_dimensions[get_column_letter(ci)].width = 12
    log.info("Excel gerado: %s", path.name)


def criar_task_clickup(evento: Evento) -> bool:
    """Cria uma task no ClickUp para um evento de alta/media materialidade."""
    if not CLICKUP_API_TOKEN or not CLICKUP_LIST_ID_ALERTAS:
        log.warning("CLICKUP_API_TOKEN ou CLICKUP_LIST_ID_ALERTAS nao definidos; pulando ClickUp")
        return False

    prioridade = 2 if evento.materialidade == "alta" else 3  # 1=urgent, 2=high, 3=normal, 4=low
    assignees = []
    if CLICKUP_ASSIGNEE_IDS:
        assignees = [int(x.strip()) for x in CLICKUP_ASSIGNEE_IDS.split(",") if x.strip()]

    # Formata data de ocorrencia (DD/MM/AAAA) e start_date em epoch ms para ClickUp.
    data_ocorrencia = ""
    start_date_ms = None
    if evento.data_entrega:
        try:
            dt = datetime.fromisoformat(evento.data_entrega.replace("Z", ""))
            data_ocorrencia = dt.strftime("%d/%m/%Y")
            start_date_ms = int(dt.timestamp() * 1000)
        except Exception:
            data_ocorrencia = evento.data_entrega[:10]

    nome_task = f"[{evento.materialidade.upper()}] {evento.empresa} - {evento.categoria}"
    if data_ocorrencia:
        nome_task += f" - {data_ocorrencia}"

    descricao = (
        f"**Data de ocorrencia:** {data_ocorrencia or '-'}\n\n"
        f"**Empresa:** {evento.empresa}\n"
        f"**CNPJ:** {evento.cnpj}\n"
        f"**Categoria:** {evento.categoria}\n"
        f"**Tipo:** {evento.tipo}\n"
        f"**Assunto:** {evento.assunto}\n"
        f"**Materialidade:** {evento.materialidade.upper()}\n"
        f"**Termos criticos:** {', '.join(evento.termos_encontrados) or '-'}\n\n"
        f"**Documento CVM:** {evento.link}\n\n"
        f"---\n"
        f"_Detectado automaticamente pelo monitor CVM Fatos Relevantes._"
    )

    payload = {
        "name": nome_task,
        "description": descricao,
        "priority": prioridade,
        "tags": ["cvm", "fato-relevante", evento.materialidade],
    }
    if assignees:
        payload["assignees"] = assignees
    if start_date_ms is not None:
        payload["start_date"] = start_date_ms
        payload["start_date_time"] = True

    url = f"https://api.clickup.com/api/v2/list/{CLICKUP_LIST_ID_ALERTAS}/task"
    headers = {"Authorization": CLICKUP_API_TOKEN, "Content-Type": "application/json"}

    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=30)
        resp.raise_for_status()
        task_id = resp.json().get("id", "?")
        log.info("Task ClickUp criada: %s (%s)", task_id, nome_task[:60])
        return True
    except Exception as exc:
        log.error("Falha ao criar task no ClickUp: %s", exc)
        return False


def salvar_last_run(eventos_novos, eventos_janela, lookback_days):
    LAST_RUN_PATH.parent.mkdir(parents=True, exist_ok=True)
    meta = {
        "executado_em": datetime.now().isoformat(timespec="seconds"),
        "eventos_novos": len(eventos_novos),
        "eventos_janela_total": eventos_janela,
        "lookback_days": lookback_days,
        "empresas_com_evento_novo": sorted({e.empresa for e in eventos_novos}),
    }
    LAST_RUN_PATH.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def main(lookback_days, dry_run):
    watchlist = carregar_watchlist()
    ipe = baixar_ipe()
    eventos = filtrar_eventos(ipe, watchlist, lookback_days)
    log.info("Total de eventos na janela: %d", len(eventos))

    if eventos:
        persistir_historico(eventos)

    seen = carregar_state()
    novos = [e for e in eventos if e.fingerprint not in seen]
    log.info("Eventos novos para alertar: %d", len(novos))

    if novos and not dry_run:
        log.info("Criando %d task(s) no ClickUp...", len(novos))
        for e in novos:
            ok = criar_task_clickup(e)
            if ok:
                seen.add(e.fingerprint)
        salvar_state(seen)
    elif novos:
        log.info("(dry-run: %d task(s) seriam criadas no ClickUp)", len(novos))

    salvar_last_run(novos, len(eventos), lookback_days)
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Monitor CVM IPE Cloud (GitHub Actions)")
    parser.add_argument("--lookback", type=int, default=DEFAULT_LOOKBACK_DAYS)
    parser.add_argument("--dry-run", action="store_true", help="Nao cria task no ClickUp")
    args = parser.parse_args()
    try:
        sys.exit(main(args.lookback, args.dry_run))
    except Exception:
        log.exception("Falha fatal")
        sys.exit(1)
